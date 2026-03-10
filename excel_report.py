import os
from collections import defaultdict
from openpyxl import Workbook
from datetime import datetime
from openpyxl.chart import BarChart, PieChart, Reference
from database.db_manager import SessionLocal
from database.models import Invoice, LineItem

OUTPUT_FOLDER = "output"

def fetch_month_data(year, month):
    if month == 12:
        next_month = datetime(year + 1, 1, 1)
    else:
        next_month = datetime(year, month + 1, 1)
    start_date = datetime(year, month, 1)
    session = SessionLocal()
    invoices = session.query(Invoice).filter(
        Invoice.created_at >= start_date,
        Invoice.created_at < next_month
    ).all()
    invoice_ids = [inv.id for inv in invoices]
    line_items = session.query(LineItem).filter(
        LineItem.invoice_id.in_(invoice_ids)
    ).all()
    session.close()
    return invoices, line_items

def generate_category_summary(invoices):
    summary = defaultdict(float)
    for invoice in invoices:
        category = invoice.category or "Other"
        summary[category] += invoice.total_amount or 0
    return summary

def generate_excel_report(year=None, month=None):
    now = datetime.now()
    year = year or now.year
    month = month or now.month
    invoices, line_items = fetch_month_data(year, month)

    if not invoices:
        print("No invoices found for this month")
        return
    
    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    file_name = f"{year}_{month:02d}_expense_report.xlsx"
    file_path = os.path.join(OUTPUT_FOLDER, file_name)
    if os.path.exists(file_path):
        os.remove(file_path)
    wb = Workbook()

    sheet1 = wb.active
    sheet1.title = "Invoices"
    headers = ["Vendor", "Invoice Number", "Invoice Date", "Due Date", "Total Amount",
        "GST Number", "Payment Status", "Category", "File Path"
    ]
    sheet1.append(headers)
    for inv in invoices:
        sheet1.append([
            inv.vendor_name,
            inv.invoice_number,
            inv.invoice_date,
            inv.due_date,
            inv.total_amount,
            inv.gst_number,
            inv.payment_status,
            inv.category,
            inv.file_path
        ])

    sheet2 = wb.create_sheet("Line Items")
    sheet2.append(["Invoice ID", "Description", "Quantity", "Unit Price", "Item Total"])
    for item in line_items:
        sheet2.append([
            item.invoice_id,
            item.description,
            item.quantity,
            item.unit_price,
            item.item_total
        ])

    sheet3 = wb.create_sheet("Category Summary")
    summary = generate_category_summary(invoices)
    sheet3.append(["Category", "Total Spend"])
    for category, amount in summary.items():
        sheet3.append([category, amount])
    
    chart = BarChart()
    chart.title = "Monthly Expense by Category"
    chart.x_axis.title = "Category"
    chart.y_axis.title = "Total Spend"
    data = Reference(sheet3, min_col=2, min_row=1, max_row=len(summary) + 1)
    categories = Reference(sheet3, min_col=1, min_row=2, max_row=len(summary) + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(categories)
    sheet3.add_chart(chart, "D2")

    pie_chart = PieChart()
    pie_chart.title = "Expense Distribution"
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(categories)
    sheet3.add_chart(pie_chart, "D20")

    wb.save(file_path)
    print("Excel report generated:", file_path)